import json
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)


def generate_excel(agent1_json: str, email_id: int, run_id: int) -> str:
    data = json.loads(agent1_json) if isinstance(agent1_json, str) else agent1_json
    filename = f"export_email{email_id}_run{run_id}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Parsed Email Data"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Email metadata section
    ws.append(["Email Metadata"])
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(bold=True, size=13)

    meta = data.get("email_metadata", {})
    for key, val in meta.items():
        ws.append([key.replace("_", " ").title(), str(val)])

    ws.append([])
    ws.append(["Classification", data.get("classification", "")])
    ws.append(["Language", data.get("language", "")])
    ws.append(["Processed At", data.get("processed_at", "")])
    ws.append([])

    # PDF content
    pdf = data.get("pdf_extracted", {})
    ws.append(["PDF Content"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=13)
    ws.append(["Title", pdf.get("title", "")])
    ws.append(["Pages", str(pdf.get("pages", ""))])
    ws.append(["OCR Confidence", str(pdf.get("ocr_confidence", ""))])
    ws.append([])

    for block in pdf.get("content_blocks", []):
        if block["type"] == "table":
            headers = block.get("headers", [])
            row_num = ws.max_row + 1
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            for row in block.get("rows", []):
                r = ws.max_row + 1
                ws.append(row)
                for col_idx in range(1, len(row) + 1):
                    ws.cell(row=r, column=col_idx).border = border
        elif block["type"] in ("header", "paragraph"):
            ws.append([block.get("text", "")])

    # Auto-width columns
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and not isinstance(cell, type(None)):
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    wb.save(filepath)
    return filename


def generate_csv(agent1_json: str, email_id: int, run_id: int) -> str:
    data = json.loads(agent1_json) if isinstance(agent1_json, str) else agent1_json
    filename = f"export_email{email_id}_run{run_id}.csv"
    filepath = os.path.join(EXPORTS_DIR, filename)

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Field", "Value"])

        meta = data.get("email_metadata", {})
        for key, val in meta.items():
            writer.writerow([key.replace("_", " ").title(), str(val)])

        writer.writerow(["Classification", data.get("classification", "")])
        writer.writerow(["Language", data.get("language", "")])
        writer.writerow([])

        pdf = data.get("pdf_extracted", {})
        writer.writerow(["PDF Title", pdf.get("title", "")])
        writer.writerow(["Pages", pdf.get("pages", "")])
        writer.writerow(["OCR Confidence", pdf.get("ocr_confidence", "")])
        writer.writerow([])

        for block in pdf.get("content_blocks", []):
            if block["type"] == "table":
                writer.writerow(block.get("headers", []))
                for row in block.get("rows", []):
                    writer.writerow(row)
            elif block["type"] in ("header", "paragraph"):
                writer.writerow([block.get("text", "")])

    return filename
